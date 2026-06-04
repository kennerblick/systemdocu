import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Server, Service, ServiceInstance, Storage

router = APIRouter(tags=["export"])

# ── Style constants ────────────────────────────────────────────────────────────

HEADER_BG   = "1E3A5F"
HDR_FONT    = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
THIN        = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED         = Side(style="medium", color="94A3B8")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

OS_HEX = {
    "linux": "3B82F6", "windows": "60A5FA",
    "proxmox": "F97316", "esxi": "22C55E",
}
SVC_HEX = {
    "postgresql": "336791", "docker": "2496ED", "kubernetes": "326CE5",
    "hyperv": "00ADEF",    "samba": "D97706",   "nfs": "B45309",    "sftp": "059669",
    "freeipa": "7C3AED",   "zabbix": "E53E3E",  "graylog": "2D3748",
    "veeam": "00B050",     "minio": "C83B0E",   "gateway": "0D9488",
    "webserver": "0EA5E9",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def _h(c: str) -> str:
    return c.lstrip("#").upper().zfill(6)


def _fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=_h(c))


def _srv_hex(srv: Server) -> str:
    if srv.environments:
        return _h(srv.environments[0].color)
    return OS_HEX.get(srv.os_type, "888888")


def _cell(ws, row: int, col: int, value, bg: str, font: Font,
          align: Alignment = LEFT) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = _fill(bg)
    c.font   = font
    c.border = CELL_BORDER
    c.alignment = align


def _header_row(ws, titles: list[str]) -> None:
    for i, t in enumerate(titles, 1):
        _cell(ws, 1, i, t, HEADER_BG, HDR_FONT, CENTER)
    ws.row_dimensions[1].height = 22


def _merge(ws, col: int, r1: int, r2: int) -> None:
    if r2 > r1:
        ws.merge_cells(
            start_row=r1, start_column=col,
            end_row=r2,   end_column=col,
        )
        c = ws.cell(row=r1, column=col)
        c.alignment = CENTER


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


DATA_FONT  = Font(color="1F2937", size=10, name="Calibri")
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
GREY_FONT  = Font(color="6B7280", size=10, name="Calibri", italic=True)

# ── Endpoint ───────────────────────────────────────────────────────────────────

@router.get("/api/export/excel")
async def export_excel(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Server).options(
            selectinload(Server.services)
            .selectinload(Service.instances)
            .selectinload(ServiceInstance.applications),
            selectinload(Server.services)
            .selectinload(Service.instances)
            .selectinload(ServiceInstance.storage),
            selectinload(Server.tags),
            selectinload(Server.environments),
            selectinload(Server.storages),
        )
    )
    servers = list(result.scalars().all())

    wb = openpyxl.Workbook()
    _build_sheet1(wb, servers)
    _build_sheet2(wb, servers)
    _build_sheet_vms(wb, servers)
    _build_sheet_db(wb, servers)
    _build_sheet_samba(wb, servers)
    _build_sheet_docker(wb, servers)
    _build_sheet_nfs(wb, servers)
    _build_sheet_storages(wb, servers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=systemdocu.xlsx"},
    )


# ── Sheet 1: Infrastruktur (Server → Service → Instanz → Anwendung) ───────────

BANDS1 = ["EFF6FF", "F0FDF4"]


def _build_sheet1(wb: openpyxl.Workbook, servers: list) -> None:
    ws = wb.active
    ws.title = "Infrastruktur"
    ws.freeze_panes = "A2"

    _header_row(ws, [
        "Server", "OS", "IP", "Umgebungen",
        "Service", "Version", "Port", "Instanz", "Anwendungen",
    ])
    _col_widths(ws, [22, 10, 16, 22, 14, 10, 7, 22, 35])

    servers = sorted(servers, key=lambda s: s.hostname.lower())
    row = 2
    for si, srv in enumerate(servers):
        envs = ", ".join(e.name for e in srv.environments)
        srv_hex = _srv_hex(srv)
        band    = BANDS1[si % 2]
        srv_r0  = row

        services = srv.services or []
        if not services:
            _write_srv_row(ws, row, srv, envs, "", "", "", "", "", band)
            row += 1
        else:
            for svc in services:
                svc_hex = SVC_HEX.get(svc.type, "4B5563")
                svc_r0  = row
                instances = svc.instances or []
                if not instances:
                    _write_srv_row(ws, row, srv, envs,
                                   svc.type, svc.version or "",
                                   str(svc.port) if svc.port else "",
                                   "", "", band)
                    _apply_svc_cell(ws, row, svc_hex)
                    row += 1
                else:
                    for inst in instances:
                        apps = ", ".join(a.name for a in inst.applications)
                        _write_srv_row(ws, row, srv, envs,
                                       svc.type, svc.version or "",
                                       str(svc.port) if svc.port else "",
                                       inst.name, apps, band)
                        _apply_svc_cell(ws, row, svc_hex)
                        row += 1

                svc_r1 = row - 1
                for c in (5, 6, 7):
                    _merge(ws, c, svc_r0, svc_r1)

        srv_r1 = row - 1
        for c in range(1, 5):
            _merge(ws, c, srv_r0, srv_r1)

        # Server name cell: coloured, bold
        top = ws.cell(row=srv_r0, column=1)
        top.fill  = _fill(srv_hex)
        top.font  = WHITE_BOLD
        top.alignment = CENTER

        # Thick left border on server group
        for r in range(srv_r0, row):
            c = ws.cell(row=r, column=1)
            c.border = Border(
                left=MED, right=THIN, top=THIN, bottom=THIN,
            )


def _write_srv_row(ws, row, srv, envs, svc_t, svc_v, svc_p, inst, apps, band):
    vals = [srv.hostname, srv.os_type, srv.ip or "", envs,
            svc_t, svc_v, svc_p, inst, apps]
    aligns = [CENTER, CENTER, CENTER, LEFT, CENTER, CENTER, CENTER, LEFT, LEFT]
    for col, (v, a) in enumerate(zip(vals, aligns), 1):
        _cell(ws, row, col, v, band, DATA_FONT, a)


def _apply_svc_cell(ws, row: int, svc_hex: str) -> None:
    c = ws.cell(row=row, column=5)
    c.fill = _fill(svc_hex)
    c.font = WHITE_BOLD
    c.alignment = CENTER


# ── Sheet 2: Anwendungen (Anwendung → Instanz → Server) ───────────────────────

BANDS2 = ["FFF7ED", "F0FDF4"]


def _build_sheet2(wb: openpyxl.Workbook, servers: list) -> None:
    ws = wb.create_sheet("Anwendungen")
    ws.freeze_panes = "A2"

    _header_row(ws, [
        "Anwendung", "Instanz", "Service", "Version",
        "Server", "OS", "IP", "Umgebungen",
    ])
    _col_widths(ws, [24, 22, 14, 10, 20, 10, 16, 22])

    # Build app → [(inst, svc, srv, envs)] map preserving app order
    app_map: dict[int, dict] = {}
    no_app: list[tuple] = []

    for srv in servers:
        envs = ", ".join(e.name for e in srv.environments)
        for svc in (srv.services or []):
            for inst in (svc.instances or []):
                if inst.applications:
                    for app in inst.applications:
                        if app.id not in app_map:
                            app_map[app.id] = {"app": app, "rows": []}
                        app_map[app.id]["rows"].append((inst, svc, srv, envs))
                else:
                    no_app.append((inst, svc, srv, envs))

    row = 2
    for ai, entry in enumerate(sorted(app_map.values(), key=lambda e: e["app"].name.lower())):
        app   = entry["app"]
        band  = BANDS2[ai % 2]
        app_r0 = row
        app_hex = _h(app.color)

        for inst, svc, srv, envs in entry["rows"]:
            svc_hex = SVC_HEX.get(svc.type, "4B5563")
            vals = [app.name, inst.name, svc.type, svc.version or "",
                    srv.hostname, srv.os_type, srv.ip or "", envs]
            aligns = [LEFT, LEFT, CENTER, CENTER, LEFT, CENTER, CENTER, LEFT]
            for col, (v, a) in enumerate(zip(vals, aligns), 1):
                _cell(ws, row, col, v, band, DATA_FONT, a)
            # App name cell: coloured
            ca = ws.cell(row=row, column=1)
            ca.fill = _fill(app_hex)
            ca.font = WHITE_BOLD
            # Service cell: coloured
            cs = ws.cell(row=row, column=3)
            cs.fill = _fill(svc_hex)
            cs.font = WHITE_BOLD
            cs.alignment = CENTER
            row += 1

        _merge(ws, 1, app_r0, row - 1)
        # Restore colour after merge reset
        top = ws.cell(row=app_r0, column=1)
        top.fill = _fill(app_hex)
        top.font = WHITE_BOLD

        # Thick left border on app group
        for r in range(app_r0, row):
            ws.cell(row=r, column=1).border = Border(
                left=MED, right=THIN, top=THIN, bottom=THIN,
            )

    # Instances with no application
    if no_app:
        no_r0 = row
        for inst, svc, srv, envs in no_app:
            svc_hex = SVC_HEX.get(svc.type, "4B5563")
            vals = ["(keine Anwendung)", inst.name, svc.type, svc.version or "",
                    srv.hostname, srv.os_type, srv.ip or "", envs]
            aligns = [LEFT, LEFT, CENTER, CENTER, LEFT, CENTER, CENTER, LEFT]
            for col, (v, a) in enumerate(zip(vals, aligns), 1):
                _cell(ws, row, col, v, "F3F4F6", GREY_FONT, a)
            row += 1
        _merge(ws, 1, no_r0, row - 1)


# ── Service-type detail sheets ─────────────────────────────────────────────────
# Each sheet shows all instances of a specific service type across all servers.
# Sheets are only added to the workbook if at least one row exists.

BANDS3 = ["FEF3C7", "ECFDF5"]


def _collect_svc_rows(servers: list, svc_types: set) -> list[dict]:
    rows = []
    for srv in sorted(servers, key=lambda s: s.hostname.lower()):
        envs = ", ".join(e.name for e in srv.environments)
        for svc in (srv.services or []):
            if svc.type in svc_types:
                for inst in sorted(svc.instances or [], key=lambda i: i.name.lower()):
                    rows.append({"srv": srv, "svc": svc, "inst": inst, "envs": envs})
    return rows


def _build_svc_instances_sheet(
    wb: openpyxl.Workbook,
    title: str,
    headers: list[str],
    col_widths: list[int],
    rows: list[dict],
    write_row,
    srv_merge_cols: list[int],
) -> None:
    """Generic sheet: one row per instance, server-info columns merged per server group."""
    if not rows:
        return
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    _header_row(ws, headers)
    _col_widths(ws, col_widths)

    row, si, i = 2, 0, 0
    while i < len(rows):
        srv    = rows[i]["srv"]
        srv_r0 = row
        band   = BANDS3[si % 2]
        si    += 1
        while i < len(rows) and rows[i]["srv"].id == srv.id:
            write_row(ws, row, rows[i], band)
            row += 1
            i   += 1
        for c in srv_merge_cols:
            _merge(ws, c, srv_r0, row - 1)
        top           = ws.cell(row=srv_r0, column=srv_merge_cols[0])
        top.fill      = _fill(_srv_hex(srv))
        top.font      = WHITE_BOLD
        top.alignment = CENTER
        for r in range(srv_r0, row):
            ws.cell(row=r, column=srv_merge_cols[0]).border = Border(
                left=MED, right=THIN, top=THIN, bottom=THIN,
            )


def _build_sheet_samba(wb: openpyxl.Workbook, servers: list) -> None:
    svc_hex = SVC_HEX["samba"]
    rows    = _collect_svc_rows(servers, {"samba"})

    def _write(ws, row, e, band):
        storage = e["inst"].storage.name if e["inst"].storage else ""
        vals   = [e["inst"].name, storage, e["srv"].hostname, e["srv"].os_type,
                  e["srv"].ip or "", e["envs"]]
        aligns = [LEFT, LEFT, CENTER, CENTER, CENTER, LEFT]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            _cell(ws, row, col, v, band, DATA_FONT, a)
        ws.cell(row=row, column=1).fill = _fill(svc_hex)
        ws.cell(row=row, column=1).font = WHITE_BOLD

    _build_svc_instances_sheet(
        wb, "Samba-Freigaben",
        ["Freigabe", "Storage", "Server", "OS", "IP", "Umgebungen"],
        [28, 20, 22, 10, 16, 22],
        rows, _write, [3, 4, 5, 6],
    )


def _build_sheet_db(wb: openpyxl.Workbook, servers: list) -> None:
    svc_hex = SVC_HEX["postgresql"]
    rows    = _collect_svc_rows(servers, {"postgresql"})

    def _write(ws, row, e, band):
        storage = e["inst"].storage.name if e["inst"].storage else ""
        vals   = [e["inst"].name, storage, e["srv"].hostname, e["srv"].os_type,
                  e["srv"].ip or "", e["envs"], e["svc"].version or ""]
        aligns = [LEFT, LEFT, CENTER, CENTER, CENTER, LEFT, CENTER]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            _cell(ws, row, col, v, band, DATA_FONT, a)
        ws.cell(row=row, column=1).fill = _fill(svc_hex)
        ws.cell(row=row, column=1).font = WHITE_BOLD

    _build_svc_instances_sheet(
        wb, "Datenbanken",
        ["Datenbank", "Storage", "Server", "OS", "IP", "Umgebungen", "PG-Version"],
        [28, 20, 22, 10, 16, 22, 12],
        rows, _write, [3, 4, 5, 6, 7],
    )


def _build_sheet_docker(wb: openpyxl.Workbook, servers: list) -> None:
    svc_hex = SVC_HEX["docker"]
    rows    = _collect_svc_rows(servers, {"docker"})

    def _write(ws, row, e, band):
        storage = e["inst"].storage.name if e["inst"].storage else ""
        vals   = [e["inst"].name, storage, e["srv"].hostname, e["srv"].os_type,
                  e["srv"].ip or "", e["envs"]]
        aligns = [LEFT, LEFT, CENTER, CENTER, CENTER, LEFT]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            _cell(ws, row, col, v, band, DATA_FONT, a)
        ws.cell(row=row, column=1).fill = _fill(svc_hex)
        ws.cell(row=row, column=1).font = WHITE_BOLD

    _build_svc_instances_sheet(
        wb, "Docker-Container",
        ["Container", "Storage", "Server", "OS", "IP", "Umgebungen"],
        [30, 20, 22, 10, 16, 22],
        rows, _write, [3, 4, 5, 6],
    )


def _build_sheet_nfs(wb: openpyxl.Workbook, servers: list) -> None:
    svc_hex = SVC_HEX["nfs"]
    rows    = _collect_svc_rows(servers, {"nfs"})

    def _write(ws, row, e, band):
        storage = e["inst"].storage.name if e["inst"].storage else ""
        vals   = [e["inst"].name, storage, e["srv"].hostname, e["srv"].os_type,
                  e["srv"].ip or "", e["envs"]]
        aligns = [LEFT, LEFT, CENTER, CENTER, CENTER, LEFT]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            _cell(ws, row, col, v, band, DATA_FONT, a)
        ws.cell(row=row, column=1).fill = _fill(svc_hex)
        ws.cell(row=row, column=1).font = WHITE_BOLD

    _build_svc_instances_sheet(
        wb, "NFS-Exporte",
        ["Export/Mount", "Storage", "Server", "OS", "IP", "Umgebungen"],
        [30, 20, 22, 10, 16, 22],
        rows, _write, [3, 4, 5, 6],
    )


def _build_sheet_vms(wb: openpyxl.Workbook, servers: list) -> None:
    """VMs sheet: Hyper-V + Proxmox combined, Typ column merged per service group."""
    rows = []
    for srv in sorted(servers, key=lambda s: s.hostname.lower()):
        envs = ", ".join(e.name for e in srv.environments)
        for svc in (srv.services or []):
            if svc.type in {"hyperv", "proxmox"}:
                for inst in sorted(svc.instances or [], key=lambda i: i.name.lower()):
                    rows.append({"srv": srv, "svc": svc, "inst": inst, "envs": envs})
    if not rows:
        return

    ws = wb.create_sheet("Virtuelle Maschinen")
    ws.freeze_panes = "A2"
    _header_row(ws, ["VM-Name", "Typ", "Storage", "Server", "OS", "IP", "Umgebungen"])
    _col_widths(ws, [30, 12, 20, 22, 10, 16, 22])

    row, si, i = 2, 0, 0
    while i < len(rows):
        srv    = rows[i]["srv"]
        srv_r0 = row
        band   = BANDS3[si % 2]
        si    += 1

        j = i
        while j < len(rows) and rows[j]["srv"].id == srv.id:
            svc       = rows[j]["svc"]
            svc_r0    = row
            typ_label = "Hyper-V" if svc.type == "hyperv" else "Proxmox"
            typ_hex   = SVC_HEX.get(svc.type, "4B5563")

            k = j
            while (k < len(rows)
                   and rows[k]["srv"].id == srv.id
                   and rows[k]["svc"].id == svc.id):
                e = rows[k]
                storage = e["inst"].storage.name if e["inst"].storage else ""
                vals   = [e["inst"].name, typ_label, storage, srv.hostname,
                          srv.os_type, srv.ip or "", e["envs"]]
                aligns = [LEFT, CENTER, LEFT, CENTER, CENTER, CENTER, LEFT]
                for col, (v, a) in enumerate(zip(vals, aligns), 1):
                    _cell(ws, row, col, v, band, DATA_FONT, a)
                ws.cell(row=row, column=1).fill = _fill(typ_hex)
                ws.cell(row=row, column=1).font = WHITE_BOLD
                row += 1
                k   += 1

            _merge(ws, 2, svc_r0, row - 1)
            ws.cell(row=svc_r0, column=2).fill      = _fill(typ_hex)
            ws.cell(row=svc_r0, column=2).font      = WHITE_BOLD
            ws.cell(row=svc_r0, column=2).alignment = CENTER
            j = k

        for c in [4, 5, 6, 7]:
            _merge(ws, c, srv_r0, row - 1)
        top           = ws.cell(row=srv_r0, column=4)
        top.fill      = _fill(_srv_hex(srv))
        top.font      = WHITE_BOLD
        top.alignment = CENTER
        for r in range(srv_r0, row):
            ws.cell(row=r, column=4).border = Border(
                left=MED, right=THIN, top=THIN, bottom=THIN,
            )
        i = j


def _build_sheet_storages(wb: openpyxl.Workbook, servers: list) -> None:
    # Build storage_id → sorted list of connected instance names from already-loaded data
    storage_instances: dict[int, list[str]] = {}
    for srv in servers:
        for svc in srv.services or []:
            for inst in svc.instances or []:
                if inst.storage_id:
                    storage_instances.setdefault(inst.storage_id, []).append(inst.name)
    for v in storage_instances.values():
        v.sort(key=str.lower)

    rows = []
    for srv in sorted(servers, key=lambda s: s.hostname.lower()):
        for st in sorted(srv.storages or [], key=lambda s: s.name.lower()):
            rows.append({"srv": srv, "st": st})
    if not rows:
        return

    ws = wb.create_sheet("Storages")
    ws.freeze_panes = "A2"
    _header_row(ws, ["Server", "OS", "IP", "Storage-Name", "RAID-Typ", "Disks", "Größe (GB)", "Services", "Beschreibung"])
    _col_widths(ws, [22, 10, 16, 24, 14, 8, 12, 36, 36])

    row, si, i = 2, 0, 0
    while i < len(rows):
        srv    = rows[i]["srv"]
        srv_r0 = row
        band   = BANDS3[si % 2]
        si    += 1
        while i < len(rows) and rows[i]["srv"].id == srv.id:
            st = rows[i]["st"]
            inst_names = ", ".join(storage_instances.get(st.id, []))
            vals   = [srv.hostname, srv.os_type, srv.ip or "",
                      st.name, st.raid_type or "", st.disk_count or "",
                      st.size_gb or "", inst_names, st.description or ""]
            aligns = [CENTER, CENTER, CENTER, LEFT, CENTER, CENTER, CENTER, LEFT, LEFT]
            for col, (v, a) in enumerate(zip(vals, aligns), 1):
                _cell(ws, row, col, v, band, DATA_FONT, a)
            row += 1
            i   += 1
        for c in [1, 2, 3]:
            _merge(ws, c, srv_r0, row - 1)
        top           = ws.cell(row=srv_r0, column=1)
        top.fill      = _fill(_srv_hex(srv))
        top.font      = WHITE_BOLD
        top.alignment = CENTER
        for r in range(srv_r0, row):
            ws.cell(row=r, column=1).border = Border(
                left=MED, right=THIN, top=THIN, bottom=THIN,
            )
