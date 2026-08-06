#!/usr/bin/env python3
"""One-time, best-effort restore of data from a systemdocu Excel export
(see app/routers/export_excel.py) into an empty/fresh database.

Reconstructs, by name-matching:
  - Servers (hostname, os_type, ip)                    [from "Infrastruktur"]
  - Environments (name) + server<->environment links    [from "Infrastruktur"]
  - Services (server, type, version, port)              [from "Infrastruktur"]
  - ServiceInstances (service, name)                     [from "Infrastruktur"]
  - Storages (server, name, raid_type, disk_count,
    size_gb, disk_size_tb, description)                  [from "Storages"]
  - instance.storage_id links                            [from "Storages"]
  - Applications (name) + server/instance links           [from "Anwendungen"]

NOT recoverable from the Excel export (will stay empty/default after this
script runs) — see the export/DB gap analysis discussed in chat:
  - internet_routers, clusters, cluster_members
  - instance_relations, relations (all dependency/connection edges)
  - servers.common_name, .fqdn, .description, .created_at, .gateway,
    .is_gateway, .gateway_router_id, .gateway_server_id
  - service_instances.fqdn, .description, .ip, .gateway, .is_gateway,
    .available, .gateway_*_id, .cluster_id
  - services.detail
  - environments/applications: .description, .color, .subnet, .gateway
  - service types with no dedicated export sheet (only recovered if they
    also appear as a row in "Infrastruktur", which always includes them)

This is a manual, one-off recovery script — not a permanent import feature.
It is safe to re-run: every entity is matched by name/relationship before
being created, so re-running it (e.g. after fixing a warning) won't create
duplicates.

Usage (inside the backend container):
    docker compose exec backend python restore_from_excel.py /path/to/systemdocu.xlsx --dry-run
    docker compose exec backend python restore_from_excel.py /path/to/systemdocu.xlsx
"""
import argparse
import asyncio
import sys

import openpyxl
from sqlalchemy import select

from app.database import AsyncSessionLocal
from app.models import (
    Server, Service, ServiceInstance, Environment, Application, Storage,
    server_environments, server_applications, instance_applications,
)


class Stats:
    def __init__(self):
        self.counts: dict[str, int] = {}

    def bump(self, key: str) -> None:
        self.counts[key] = self.counts.get(key, 0) + 1


def load_grid(ws) -> list[list]:
    """Sheet contents as a 2D list with merged-cell values forward-filled,
    so every logical row can be read independently of the export's visual
    row-spanning merges."""
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    for rng in ws.merged_cells.ranges:
        anchor = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = anchor
    return grid


def split_names(cell) -> list[str]:
    if not cell:
        return []
    return [n.strip() for n in str(cell).split(",") if n.strip()]


def s(cell) -> str | None:
    if cell is None:
        return None
    text = str(cell).strip()
    return text or None


def as_int(cell):
    try:
        return int(cell)
    except (TypeError, ValueError):
        return None


async def get_or_create_server(db, cache, hostname, os_type, ip, stats) -> Server:
    hostname = hostname.strip()
    if hostname in cache:
        srv = cache[hostname]
    else:
        res = await db.execute(select(Server).where(Server.hostname == hostname))
        srv = res.scalar_one_or_none()
        if srv is None:
            srv = Server(hostname=hostname, os_type=os_type or "linux", ip=ip)
            db.add(srv)
            await db.flush()
            stats.bump("servers_created")
            cache[hostname] = srv
            return srv
        cache[hostname] = srv
    if os_type and not srv.os_type:
        srv.os_type = os_type
    if ip and not srv.ip:
        srv.ip = ip
    return srv


async def get_or_create_environment(db, cache, name, stats) -> Environment:
    name = name.strip()
    if name in cache:
        return cache[name]
    res = await db.execute(select(Environment).where(Environment.name == name))
    env = res.scalar_one_or_none()
    if env is None:
        env = Environment(name=name)
        db.add(env)
        await db.flush()
        stats.bump("environments_created")
    cache[name] = env
    return env


async def link_server_environment(db, srv, env, seen, stats) -> None:
    key = (srv.id, env.id)
    if key in seen:
        return
    seen.add(key)
    res = await db.execute(
        select(server_environments).where(
            server_environments.c.server_id == srv.id,
            server_environments.c.environment_id == env.id,
        )
    )
    if res.first() is None:
        await db.execute(server_environments.insert().values(server_id=srv.id, environment_id=env.id))
        stats.bump("server_environment_links_created")


async def get_or_create_service(db, cache, srv, svc_type, version, port, stats) -> Service:
    key = (srv.id, svc_type)
    svc = cache.get(key)
    if svc is None:
        res = await db.execute(select(Service).where(Service.server_id == srv.id, Service.type == svc_type))
        svc = res.scalars().first()
        if svc is None:
            svc = Service(server_id=srv.id, type=svc_type)
            db.add(svc)
            await db.flush()
            stats.bump("services_created")
        cache[key] = svc
    if version and not svc.version:
        svc.version = version
    if port is not None and not svc.port:
        svc.port = port
    return svc


async def get_or_create_instance(db, cache, svc, name, stats) -> ServiceInstance:
    key = (svc.id, name)
    inst = cache.get(key)
    if inst is not None:
        return inst
    res = await db.execute(
        select(ServiceInstance).where(ServiceInstance.service_id == svc.id, ServiceInstance.name == name)
    )
    inst = res.scalars().first()
    if inst is None:
        inst = ServiceInstance(service_id=svc.id, name=name)
        db.add(inst)
        await db.flush()
        stats.bump("instances_created")
    cache[key] = inst
    return inst


async def import_infrastruktur(db, wb, stats):
    """Reconstructs Server / Environment / Service / ServiceInstance from the
    'Infrastruktur' sheet — the only sheet that lists every server/service/
    instance regardless of service type."""
    server_cache: dict[str, Server] = {}
    instance_index: dict[tuple, ServiceInstance] = {}

    if "Infrastruktur" not in wb.sheetnames:
        print("WARN: sheet 'Infrastruktur' not found — nothing to restore.", file=sys.stderr)
        return server_cache, instance_index

    grid = load_grid(wb["Infrastruktur"])
    env_cache: dict[str, Environment] = {}
    server_env_seen: set[tuple] = set()
    service_cache: dict[tuple, Service] = {}
    instance_cache: dict[tuple, ServiceInstance] = {}

    for row in grid[1:]:
        hostname = s(row[0]) if len(row) > 0 else None
        if not hostname:
            continue
        os_type = s(row[1]) if len(row) > 1 else None
        ip = s(row[2]) if len(row) > 2 else None
        envs_str = row[3] if len(row) > 3 else None
        svc_type = s(row[4]) if len(row) > 4 else None
        version = s(row[5]) if len(row) > 5 else None
        port = as_int(row[6]) if len(row) > 6 else None
        inst_name = s(row[7]) if len(row) > 7 else None

        srv = await get_or_create_server(db, server_cache, hostname, os_type, ip, stats)

        for env_name in split_names(envs_str):
            env = await get_or_create_environment(db, env_cache, env_name, stats)
            await link_server_environment(db, srv, env, server_env_seen, stats)

        if svc_type:
            svc = await get_or_create_service(db, service_cache, srv, svc_type, version, port, stats)
            if inst_name:
                inst = await get_or_create_instance(db, instance_cache, svc, inst_name, stats)
                instance_index[(hostname, svc_type, inst_name)] = inst

    return server_cache, instance_index


async def import_storages(db, wb, server_cache, instance_by_host, stats):
    """Reconstructs Storage rows and best-effort instance.storage_id links
    from the 'Storages' sheet (skipped if the sheet doesn't exist — happens
    when the source system had no storages at export time)."""
    if "Storages" not in wb.sheetnames:
        return

    grid = load_grid(wb["Storages"])
    storage_cache: dict[tuple, Storage] = {}

    for row in grid[1:]:
        hostname = s(row[0]) if len(row) > 0 else None
        if not hostname:
            continue
        storage_name = s(row[3]) if len(row) > 3 else None
        if not storage_name:
            continue
        raid_type = s(row[4]) if len(row) > 4 else None
        disk_count = as_int(row[5]) if len(row) > 5 else None
        size_gb = as_int(row[6]) if len(row) > 6 else None
        disk_size_tb = as_int(row[7]) if len(row) > 7 else None
        inst_names_str = row[8] if len(row) > 8 else None
        description = s(row[9]) if len(row) > 9 else None

        srv = await get_or_create_server(db, server_cache, hostname, None, None, stats)

        key = (srv.id, storage_name)
        storage = storage_cache.get(key)
        if storage is None:
            res = await db.execute(select(Storage).where(Storage.server_id == srv.id, Storage.name == storage_name))
            storage = res.scalars().first()
            if storage is None:
                storage = Storage(server_id=srv.id, name=storage_name)
                db.add(storage)
                await db.flush()
                stats.bump("storages_created")
            storage_cache[key] = storage

        if raid_type and not storage.raid_type:
            storage.raid_type = raid_type
        if disk_count is not None and not storage.disk_count:
            storage.disk_count = disk_count
        if size_gb is not None and not storage.size_gb:
            storage.size_gb = size_gb
        if disk_size_tb is not None and not storage.disk_size_tb:
            storage.disk_size_tb = disk_size_tb
        if description and not storage.description:
            storage.description = description

        for inst_name in split_names(inst_names_str):
            for inst in instance_by_host.get((hostname, inst_name), []):
                if inst.storage_id is None:
                    inst.storage_id = storage.id
                    stats.bump("instance_storage_links_set")


async def import_applications(db, wb, server_cache, instance_index, instance_by_host, stats):
    """Reconstructs Application rows and server-level/instance-level links
    from the 'Anwendungen' sheet, which (unlike 'Infrastruktur') distinguishes
    a whole-server assignment ("— ganzer Server —") from an instance one."""
    if "Anwendungen" not in wb.sheetnames:
        return

    grid = load_grid(wb["Anwendungen"])
    app_cache: dict[str, Application] = {}
    server_app_seen: set[tuple] = set()
    instance_app_seen: set[tuple] = set()

    for row in grid[1:]:
        app_name = s(row[0]) if len(row) > 0 else None
        if not app_name or app_name == "(keine Anwendung)":
            continue
        inst_label = s(row[1]) if len(row) > 1 else None
        svc_type = s(row[2]) if len(row) > 2 else None
        hostname = s(row[4]) if len(row) > 4 else None
        if not hostname:
            continue

        if app_name in app_cache:
            app = app_cache[app_name]
        else:
            res = await db.execute(select(Application).where(Application.name == app_name))
            app = res.scalar_one_or_none()
            if app is None:
                app = Application(name=app_name)
                db.add(app)
                await db.flush()
                stats.bump("applications_created")
            app_cache[app_name] = app

        srv = server_cache.get(hostname)
        if srv is None:
            res = await db.execute(select(Server).where(Server.hostname == hostname))
            srv = res.scalar_one_or_none()
            if srv is None:
                print(f"WARN: unknown server '{hostname}' referenced for application "
                      f"'{app_name}', skipped", file=sys.stderr)
                stats.bump("application_links_skipped")
                continue
            server_cache[hostname] = srv

        if not inst_label or inst_label == "— ganzer Server —":
            key = (srv.id, app.id)
            if key not in server_app_seen:
                server_app_seen.add(key)
                exists = await db.execute(
                    select(server_applications).where(
                        server_applications.c.server_id == srv.id,
                        server_applications.c.application_id == app.id,
                    )
                )
                if exists.first() is None:
                    await db.execute(server_applications.insert().values(server_id=srv.id, application_id=app.id))
                    stats.bump("server_application_links_created")
            continue

        inst = instance_index.get((hostname, svc_type, inst_label)) if svc_type else None
        if inst is None:
            candidates = instance_by_host.get((hostname, inst_label), [])
            if len(candidates) == 1:
                inst = candidates[0]
            elif len(candidates) > 1:
                inst = candidates[0]
                print(f"WARN: ambiguous instance '{inst_label}' on '{hostname}' for application "
                      f"'{app_name}' ({len(candidates)} matches) — used the first one, "
                      f"double-check in the UI", file=sys.stderr)
        if inst is None:
            print(f"WARN: could not find instance '{inst_label}' on '{hostname}' for application "
                  f"'{app_name}', skipped", file=sys.stderr)
            stats.bump("application_links_skipped")
            continue

        key = (inst.id, app.id)
        if key not in instance_app_seen:
            instance_app_seen.add(key)
            exists = await db.execute(
                select(instance_applications).where(
                    instance_applications.c.instance_id == inst.id,
                    instance_applications.c.application_id == app.id,
                )
            )
            if exists.first() is None:
                await db.execute(instance_applications.insert().values(instance_id=inst.id, application_id=app.id))
                stats.bump("instance_application_links_created")


async def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("xlsx_path", help="Path to a systemdocu Excel export (systemdocu.xlsx)")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse and report what would happen, but roll back instead of committing",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    stats = Stats()

    async with AsyncSessionLocal() as db:
        server_cache, instance_index = await import_infrastruktur(db, wb, stats)

        instance_by_host: dict[tuple, list] = {}
        for (hostname, _svc_type, inst_name), inst in instance_index.items():
            instance_by_host.setdefault((hostname, inst_name), []).append(inst)

        await import_storages(db, wb, server_cache, instance_by_host, stats)
        await import_applications(db, wb, server_cache, instance_index, instance_by_host, stats)

        if args.dry_run:
            await db.rollback()
            print("--- DRY RUN: no changes were committed ---")
        else:
            await db.commit()
            print("--- Changes committed ---")

    print("\nSummary:")
    for key in sorted(stats.counts):
        print(f"  {key}: {stats.counts[key]}")
    if not stats.counts:
        print("  (nothing found to restore)")

    print(
        "\nNOT recoverable from the Excel export (unaffected by this run — see script "
        "docstring for the full list): internet routers, clusters, all server/instance "
        "relations, server description/common_name/fqdn/created_at/gateway fields, "
        "service detail text, environment/application description+color+subnet. "
        "Review the restored data in the UI before relying on it."
    )


if __name__ == "__main__":
    asyncio.run(main())
